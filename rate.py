#!/usr/bin/env python
# coding: utf-8

# In[3]:



from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer 
 

def sentiment_scores(sentence): 
 
    
   sid_obj = SentimentIntensityAnalyzer() 
 
   
   sentiment_dict = sid_obj.polarity_scores(sentence) 
   
   

   return min(4,int(6*abs(sentiment_dict['compound'])))

   


# In[4]:


print(sentiment_scores("The sound was very awesome. The best bass and smooth sound. There is no noise like sound. So I like this product very much"))


# In[5]:



    
import openpyxl 
import xlrd 
  
# Give the location of the file 
path = r"I:\extras\internship\KNIT\review (1).xlsx"
  
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
  
sheet_obj = wb_obj.active 

m_row = sheet_obj.max_row
  
max_col = sheet_obj.max_column 
  

  
  
wb = xlrd.open_workbook(path) 

sheet = wb.sheet_by_index(0) 
  
sheet.cell_value(0, 0) 
  
for i in range(1,m_row):
    l = sheet.row_values(i)
    a1 = sentiment_scores(l[0])
    a2=0
    a3=0
    if int(l[1]) == 0:
        a2=1
    if int(l[2]) == 1:
        a3=2
    a4=min(10,l[3])
        
    net_sum = 2*(a1+a2+a3) + (40*a4//100)  
    print("{} : {:.2f}".format(l[0],(net_sum/18)*100))
    print()
    



# In[ ]:




